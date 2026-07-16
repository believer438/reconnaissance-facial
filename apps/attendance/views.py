from __future__ import annotations

import base64
import csv
import io
import json
from datetime import timedelta

from django.contrib import messages
from django.core.files.base import ContentFile
from django.core.paginator import Paginator
from django.db.models import Count, Q
from django.http import HttpRequest, HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST

import datetime

import time as _time_module

from django.contrib.auth import get_user_model

from .forms import (
    AttendanceManualEditForm,
    CameraForm,
    ClasseForm,
    ClassroomScheduleForm,
    CourseForm,
    CourseSessionForm,
    DailyAttendanceEditForm,
    EnrollmentForm,
    JourFerieForm,
    ReviewQueueValidateForm,
    SalleForm,
    ScheduleForm,
    SchoolDayConfigForm,
    SessionCancelForm,
    StudentForm,
    SystemConfigForm,
    UserCreateForm,
    UserEditForm,
    UserResetPasswordForm,
)
from .models import (
    AttendanceAuditLog,
    AttendanceRecord,
    Camera,
    Classe,
    ClassroomSchedule,
    Course,
    CourseSession,
    DailyAttendance,
    Enrollment,
    FACULTY_CHOICES,
    FaceDetectionEvent,
    FaceEmbedding,
    JourFerie,
    RecognitionReviewQueue,
    Salle,
    Schedule,
    SchoolDayConfig,
    Student,
    SystemConfig,
    SystemLog,
    TrainingHistory,
    TrainingPhoto,
    UnknownFaceLog,
)
from .services.paths import LABEL_FILE, MODEL_FILE
from .services.recognition import (
    recognize_from_image_bytes,
    save_attendance_from_results,
)
from .services.daily_attendance import (
    absence_generation_block_reason,
    arrival_status,
    day_block_reason,
    generate_absences_for_date,
    register_check_in,
    time_block_reason,
)
from .services.training import train_model


def _model_ready() -> bool:
    return MODEL_FILE.exists() and LABEL_FILE.exists()


# ─────────────────────────────────────────────────────────────────────────────
# HELPER JOURNALISATION — _syslog() à appeler depuis toutes les vues clés
# ─────────────────────────────────────────────────────────────────────────────

def _syslog(
    request: HttpRequest,
    action: str,
    object_type: str = "",
    object_id: int | None = None,
    object_repr: str = "",
    details: str = "",
    success: bool = True,
) -> None:
    """
    Enregistre une action dans SystemLog.
    Non-bloquant : toute exception est silencieusement ignorée.
    """
    try:
        ip = request.META.get("HTTP_X_FORWARDED_FOR", request.META.get("REMOTE_ADDR", ""))
        if ip and "," in ip:
            ip = ip.split(",")[0].strip()
        user = request.user.username if request.user.is_authenticated else "anonyme"
        SystemLog.objects.create(
            user=user,
            action=action,
            object_type=object_type,
            object_id=object_id,
            object_repr=object_repr[:255],
            details=details[:1000],
            ip_address=ip[:45],
            success=success,
        )
    except Exception:
        pass


# ─────────────────────────────────────────────────────────────────────────────
# DASHBOARD
# ─────────────────────────────────────────────────────────────────────────────

def dashboard(request: HttpRequest) -> HttpResponse:
    ready = _model_ready()
    today = timezone.localdate()
    daily_records = DailyAttendance.objects.filter(date=today).select_related(
        "student__classe", "camera_entree"
    )
    day_config = SchoolDayConfig.get()
    daily_block_reason = _daily_day_block_reason(today, day_config)
    total_students = Student.objects.filter(is_active=True).count()
    untrained_photos = TrainingPhoto.objects.filter(trained_at__isnull=True).count()

    seuil_absences = SystemConfig.get().seuil_alerte_absences
    students_en_alerte = (
        Student.objects.filter(is_active=True)
        .annotate(nb_absences=Count("daily_attendances", filter=Q(daily_attendances__status=DailyAttendance.STATUS_ABSENT)))
        .filter(nb_absences__gte=seuil_absences)
        .select_related("classe")
        .order_by("-nb_absences")[:10]
    )

    prochain_ferie = JourFerie.objects.filter(date__gte=today).first()
    auj_ferie = JourFerie.objects.filter(date=today).first()
    embedding_total = FaceEmbedding.objects.filter(student__is_active=True).count()

    daily_present = daily_records.filter(status=DailyAttendance.STATUS_PRESENT).count()
    daily_retard = daily_records.filter(status=DailyAttendance.STATUS_RETARD).count()
    daily_absent = daily_records.filter(status=DailyAttendance.STATUS_ABSENT).count()
    daily_excuse = daily_records.filter(status=DailyAttendance.STATUS_EXCUSE).count()
    daily_non_enregistres = 0 if daily_block_reason else max(0, total_students - daily_records.count())

    context = {
        "student_total": total_students,
        "classe_total": Classe.objects.filter(is_active=True).count(),
        "salle_total": Salle.objects.filter(is_active=True).count(),
        "photo_total": TrainingPhoto.objects.count(),
        "untrained_photos": untrained_photos,
        "embedding_total": embedding_total,
        "attendance_total": DailyAttendance.objects.count(),
        "camera_total": Camera.objects.filter(is_active=True).count(),
        "today_total": daily_records.count(),
        "today_present": daily_present,
        "today_late": daily_retard,
        "today_absent": daily_absent,
        "today_excuse": daily_excuse,
        "unknown_total": UnknownFaceLog.objects.count(),
        "model_ready": ready,
        "model_path": str(MODEL_FILE) if ready else None,
        "recent_unknowns": UnknownFaceLog.objects.all()[:4],
        "students_en_alerte": students_en_alerte,
        "seuil_absences": seuil_absences,
        "prochain_ferie": prochain_ferie,
        "auj_ferie": auj_ferie,
        "day_config": day_config,
        "daily_non_enregistres": daily_non_enregistres,
        "daily_recent_arrivals": daily_records.filter(heure_entree__isnull=False).order_by("-heure_entree")[:10],
        "daily_check_in_cameras": Camera.objects.filter(is_active=True, zone_type=Camera.ZONE_CHECK_IN),
        "daily_taux_presence": round((daily_present + daily_retard + daily_excuse) / total_students * 100, 1) if total_students else 0,
        "daily_block_reason": daily_block_reason,
    }
    return render(request, "attendance/dashboard.html", context)


# ─────────────────────────────────────────────────────────────────────────────
# CLASSES
# ─────────────────────────────────────────────────────────────────────────────

def classe_list(request: HttpRequest) -> HttpResponse:
    if request.method == "POST":
        form = ClasseForm(request.POST)
        if form.is_valid():
            classe = form.save()
            messages.success(request, f"Classe '{classe}' creee.")
            return redirect("attendance:classe_list")
        messages.error(request, "Erreur dans le formulaire.")
    else:
        form = ClasseForm()
    classes = Classe.objects.annotate(
        nb_eleves=Count("students", filter=Q(students__is_active=True)),
        nb_horaires=Count("schedules"),
    ).order_by("niveau", "nom")
    return render(request, "attendance/classe_list.html", {"classes": classes, "form": form})


def classe_detail(request: HttpRequest, classe_id: int) -> HttpResponse:
    """Détail d'une classe: élèves, présence journalière et derniers mouvements."""
    classe = get_object_or_404(Classe, id=classe_id)
    date_str = request.GET.get("date", timezone.localdate().isoformat())
    try:
        target_date = datetime.date.fromisoformat(date_str)
    except ValueError:
        target_date = timezone.localdate()

    students = (
        Student.objects.filter(classe=classe, is_active=True)
        .annotate(
            photo_total=Count("photos", distinct=True),
            embedding_total=Count("face_embeddings", distinct=True),
        )
        .order_by("full_name")
    )
    student_ids = list(students.values_list("id", flat=True))
    daily_records = {
        record.student_id: record
        for record in DailyAttendance.objects.filter(
            student_id__in=student_ids,
            date=target_date,
        ).select_related("student", "camera_entree")
    }
    day_config = SchoolDayConfig.get()
    absence_block = _daily_absence_generation_block_reason(
        target_date,
        timezone.localtime().time().replace(tzinfo=None, microsecond=0),
        day_config,
    )

    rows = [
        {
            "student": student,
            "record": daily_records.get(student.id),
        }
        for student in students
    ]

    present_like = {
        DailyAttendance.STATUS_PRESENT,
        DailyAttendance.STATUS_RETARD,
        DailyAttendance.STATUS_EXCUSE,
    }
    nb_students = len(rows)
    nb_present = sum(1 for row in rows if row["record"] and row["record"].status in present_like)
    nb_retard = sum(1 for row in rows if row["record"] and row["record"].status == DailyAttendance.STATUS_RETARD)
    nb_absent = sum(
        1 for row in rows
        if row["record"] and row["record"].status == DailyAttendance.STATUS_ABSENT
    )
    if not absence_block:
        nb_absent += sum(1 for row in rows if row["record"] is None)
    nb_excuse = sum(1 for row in rows if row["record"] and row["record"].status == DailyAttendance.STATUS_EXCUSE)
    taux_presence = round(nb_present / nb_students * 100, 1) if nb_students else 0.0

    recent_daily_records = (
        DailyAttendance.objects.filter(student__classe=classe)
        .select_related("student", "camera_entree")
        .order_by("-date", "-updated_at")[:20]
    )

    return render(request, "attendance/classe_detail.html", {
        "classe": classe,
        "target_date": target_date,
        "prev_date": (target_date - datetime.timedelta(days=1)).isoformat(),
        "next_date": (target_date + datetime.timedelta(days=1)).isoformat(),
        "rows": rows,
        "recent_daily_records": recent_daily_records,
        "nb_students": nb_students,
        "nb_present": nb_present,
        "nb_retard": nb_retard,
        "nb_absent": nb_absent,
        "nb_excuse": nb_excuse,
        "taux_presence": taux_presence,
        "absence_block": absence_block,
    })


def classe_delete(request: HttpRequest, classe_id: int) -> HttpResponse:
    classe = get_object_or_404(Classe, id=classe_id)
    if request.method == "POST":
        name = str(classe)
        classe.delete()
        messages.success(request, f"Classe '{name}' supprimee.")
        return redirect("attendance:classe_list")
    return render(request, "attendance/classe_confirm_delete.html", {"classe": classe})


# ─────────────────────────────────────────────────────────────────────────────
# SALLES
# ─────────────────────────────────────────────────────────────────────────────

def salle_list(request: HttpRequest) -> HttpResponse:
    if request.method == "POST":
        form = SalleForm(request.POST)
        if form.is_valid():
            salle = form.save()
            messages.success(request, f"Salle '{salle}' ajoutee.")
            return redirect("attendance:salle_list")
        messages.error(request, "Erreur dans le formulaire.")
    else:
        form = SalleForm()
    salles = Salle.objects.annotate(
        nb_cameras=Count("cameras"),
        nb_horaires=Count("schedules"),
    ).order_by("batiment", "nom")
    return render(request, "attendance/salle_list.html", {"salles": salles, "form": form})


def salle_delete(request: HttpRequest, salle_id: int) -> HttpResponse:
    salle = get_object_or_404(Salle, id=salle_id)
    if request.method == "POST":
        name = str(salle)
        salle.delete()
        messages.success(request, f"Salle '{name}' supprimee.")
        return redirect("attendance:salle_list")
    return render(request, "attendance/salle_confirm_delete.html", {"salle": salle})


# ─────────────────────────────────────────────────────────────────────────────
# JOURS FERIES
# ─────────────────────────────────────────────────────────────────────────────

def jours_feries_list(request: HttpRequest) -> HttpResponse:
    if request.method == "POST":
        form = JourFerieForm(request.POST)
        if form.is_valid():
            jour = form.save()
            messages.success(request, f"'{jour.nom}' ajoute au calendrier.")
            return redirect("attendance:jours_feries_list")
        messages.error(request, "Erreur dans le formulaire.")
    else:
        form = JourFerieForm()
    today = timezone.localdate()
    jours = JourFerie.objects.all()
    jours_passes = jours.filter(date__lt=today)
    jours_futurs = jours.filter(date__gte=today)
    return render(request, "attendance/jours_feries_list.html", {
        "form": form,
        "jours_futurs": jours_futurs,
        "jours_passes": jours_passes,
        "today": today,
    })


def jours_feries_delete(request: HttpRequest, jour_id: int) -> HttpResponse:
    jour = get_object_or_404(JourFerie, id=jour_id)
    if request.method == "POST":
        nom = jour.nom
        jour.delete()
        messages.success(request, f"'{nom}' supprime du calendrier.")
    return redirect("attendance:jours_feries_list")


# ─────────────────────────────────────────────────────────────────────────────
# STUDENTS
# ─────────────────────────────────────────────────────────────────────────────

def student_list(request: HttpRequest) -> HttpResponse:
    q = request.GET.get("q", "").strip()
    faculty = request.GET.get("faculty", "").strip()
    classe_id = request.GET.get("classe", "").strip()
    qs = Student.objects.select_related("classe").annotate(photo_total=Count("photos")).order_by("full_name")
    if q:
        qs = qs.filter(Q(full_name__icontains=q) | Q(student_code__icontains=q))
    if faculty:
        qs = qs.filter(faculty=faculty)
    if classe_id:
        qs = qs.filter(classe_id=classe_id)
    return render(request, "attendance/student_list.html", {
        "students": qs, "q": q, "faculty_filter": faculty,
        "faculty_choices": FACULTY_CHOICES,
        "classes": Classe.objects.filter(is_active=True).order_by("niveau", "nom"),
        "classe_filter": classe_id,
    })


def api_student_code_check(request: HttpRequest) -> JsonResponse:
    code = request.GET.get("code", "").strip().upper()
    exclude_id = request.GET.get("exclude", "").strip()
    if not code:
        return JsonResponse({"available": False, "message": "Matricule vide."})
    qs = Student.objects.filter(student_code__iexact=code)
    if exclude_id.isdigit():
        qs = qs.exclude(pk=int(exclude_id))
    available = not qs.exists()
    return JsonResponse({
        "available": available,
        "code": code,
        "message": "Disponible" if available else "Deja utilise",
    })


def _save_photo_with_embedding(student, file, angle_tag: str = "") -> tuple[int, int]:
    """Sauvegarde une photo d'entraînement et génère son embedding en un seul appel."""
    from .services import sface
    from .services.embedding import extract_embedding_from_image, vector_to_bytes
    from .models import FaceEmbedding
    photo = TrainingPhoto.objects.create(
        student=student,
        image=file,
        angle_tag=angle_tag or "",   # jamais None — le champ n'accepte pas NULL
    )
    try:
        image_bytes = photo.image.read()
        if sface.models_available():
            embedding, quality = sface.extract_largest_feature(image_bytes)
            encoded = sface.vector_to_bytes(embedding) if embedding is not None else None
        else:
            # extract_embedding_from_image retourne (ndarray|None, blur_score, quality)
            embedding, blur_s, quality = extract_embedding_from_image(image_bytes)
            encoded = vector_to_bytes(embedding) if embedding is not None else None
        if embedding is not None:
            FaceEmbedding.objects.create(
                student=student,
                photo=photo,
                vector=encoded,
                score_qualite=quality,
                angle_tag=angle_tag or "",
            )
            TrainingPhoto.objects.filter(pk=photo.pk).update(trained_at=timezone.now(), face_detected=True)
            return 1, 1
    except Exception:
        pass
    TrainingPhoto.objects.filter(pk=photo.pk).update(face_detected=False)
    return 1, 0


def student_create(request: HttpRequest) -> HttpResponse:
    if request.method == "POST":
        form = StudentForm(request.POST, request.FILES)
        if form.is_valid():
            student = form.save(commit=False)
            if student.classe:
                student.classroom = str(student.classe)
            student.save()
            nb_photos = 0
            nb_embed = 0
            for file in request.FILES.getlist("photos"):
                p, e = _save_photo_with_embedding(student, file)
                nb_photos += p; nb_embed += e
            for file in request.FILES.getlist("photos_webcam"):
                p, e = _save_photo_with_embedding(student, file, "webcam")
                nb_photos += p; nb_embed += e
            _syslog(request, SystemLog.ACTION_CREATE, SystemLog.OBJ_STUDENT,
                    student.id, student.full_name,
                    f"Code: {student.student_code} | Classe: {student.classroom} | Photos: {nb_photos} | Empreintes: {nb_embed}")
            messages.success(request, f"{student.full_name} inscrit avec succès ({nb_embed} empreinte(s) faciale(s) enregistrée(s)).")
            return redirect("attendance:student_detail", student_id=student.id)
        _syslog(request, SystemLog.ACTION_CREATE, SystemLog.OBJ_STUDENT,
                details=f"ECHEC creation: {form.errors.as_text()[:200]}", success=False)
        messages.error(request, f"Erreur : {form.errors.as_text()}")
    else:
        form = StudentForm()
    return render(request, "attendance/student_create.html", {"form": form})


def student_detail(request: HttpRequest, student_id: int) -> HttpResponse:
    student = get_object_or_404(
        Student.objects.select_related("classe").prefetch_related("photos"),
        id=student_id,
    )
    records = student.daily_attendances.select_related("camera_entree").order_by("-date", "-updated_at")[:30]
    trained = student.photos.filter(trained_at__isnull=False).count()
    untrained = student.photos.filter(trained_at__isnull=True).count()
    all_daily = student.daily_attendances.all()
    presents = all_daily.filter(status=DailyAttendance.STATUS_PRESENT).count()
    retards = all_daily.filter(status=DailyAttendance.STATUS_RETARD).count()
    absences = all_daily.filter(status=DailyAttendance.STATUS_ABSENT).count()
    excuses = all_daily.filter(status=DailyAttendance.STATUS_EXCUSE).count()
    total_daily = all_daily.count()
    return render(request, "attendance/student_detail.html", {
        "student": student,
        "records": records,
        "attendance_rate": student.attendance_rate(),
        "trained_photos": trained,
        "untrained_photos": untrained,
        "presences_count": presents,
        "retards_count": retards,
        "absences_count": absences,
        "excuses_count": excuses,
        "total_daily_count": total_daily,
    })


def student_edit(request: HttpRequest, student_id: int) -> HttpResponse:
    student = get_object_or_404(Student, id=student_id)
    if request.method == "POST":
        form = StudentForm(request.POST, request.FILES, instance=student)
        if form.is_valid():
            s = form.save(commit=False)
            if s.classe:
                s.classroom = str(s.classe)
            s.save()
            nb_photos = 0
            nb_embed = 0
            for file in request.FILES.getlist("photos"):
                p, e = _save_photo_with_embedding(student, file)
                nb_photos += p; nb_embed += e
            details = f"Code: {s.student_code} | Classe: {s.classroom}"
            if nb_photos:
                details += f" | +{nb_photos} photo(s) | +{nb_embed} empreinte(s)"
            _syslog(request, SystemLog.ACTION_UPDATE, SystemLog.OBJ_STUDENT,
                    student.id, student.full_name, details)
            messages.success(request, "Modifications enregistrées.")
            return redirect("attendance:student_detail", student_id=student.id)
        messages.error(request, "Erreur dans le formulaire.")
    else:
        form = StudentForm(instance=student)
    return render(request, "attendance/student_edit.html", {"form": form, "student": student})


def student_delete(request: HttpRequest, student_id: int) -> HttpResponse:
    student = get_object_or_404(Student, id=student_id)
    if request.method == "POST":
        name = student.full_name
        code = student.student_code
        student.delete()
        _syslog(request, SystemLog.ACTION_DELETE, SystemLog.OBJ_STUDENT,
                student_id, name, f"Code: {code}")
        messages.success(request, f"{name} supprime.")
        return redirect("attendance:student_list")
    return render(request, "attendance/student_confirm_delete.html", {"student": student})


def student_add_photos(request: HttpRequest, student_id: int) -> HttpResponse:
    student = get_object_or_404(Student, id=student_id)
    if request.method == "POST":
        files = request.FILES.getlist("photos")
        angle_tag = request.POST.get("angle_tag", "") or ""   # jamais None
        if not files:
            messages.error(request, "Aucun fichier sélectionné.")
        else:
            nb_ok = 0
            nb_embed = 0
            for file in files:
                p, e = _save_photo_with_embedding(student, file, angle_tag)
                nb_ok += p
                nb_embed += e
            details = f"{nb_ok} photo(s) ajoutée(s), {nb_embed} empreinte(s) générée(s)"
            _syslog(request, SystemLog.ACTION_CREATE, SystemLog.OBJ_PHOTO,
                    student.id, student.full_name, details)
            if nb_embed > 0:
                messages.success(request, f"{nb_ok} photo(s) ajoutée(s) — {nb_embed} empreinte(s) faciale(s) enregistrée(s). Reconnaissance disponible immédiatement.")
            else:
                messages.warning(request, f"{nb_ok} photo(s) ajoutée(s). Aucun visage détecté — vérifiez la qualité des photos.")
    return redirect("attendance:student_detail", student_id=student.id)


def photo_delete(request: HttpRequest, photo_id: int) -> HttpResponse:
    photo = get_object_or_404(TrainingPhoto, id=photo_id)
    student_id = photo.student_id
    if request.method == "POST":
        photo.delete()
        messages.success(request, "Photo supprimee.")
    return redirect("attendance:student_detail", student_id=student_id)


# ─────────────────────────────────────────────────────────────────────────────
# COURSES
# ─────────────────────────────────────────────────────────────────────────────

def course_list(request: HttpRequest) -> HttpResponse:
    if request.method == "POST":
        form = CourseForm(request.POST)
        if form.is_valid():
            course = form.save()
            messages.success(request, f"Cours {course.code} cree.")
            return redirect("attendance:course_detail", course_id=course.id)
        messages.error(request, "Erreur dans le formulaire.")
    else:
        form = CourseForm()
    courses = Course.objects.annotate(
        student_count=Count("enrollments", distinct=True),
        session_count=Count("sessions", distinct=True),
    )
    return render(request, "attendance/course_list.html", {"courses": courses, "form": form})


def course_detail(request: HttpRequest, course_id: int) -> HttpResponse:
    course = get_object_or_404(Course, id=course_id)
    enrollments = Enrollment.objects.filter(course=course).select_related("student__classe")
    sessions = CourseSession.objects.filter(course=course).annotate(
        present_count=Count("attendance_records", filter=Q(attendance_records__status__in=["present", "late"]))
    ).select_related("schedule__classe")
    enrolled_ids = enrollments.values_list("student_id", flat=True)
    enrollment_form = EnrollmentForm()
    enrollment_form.fields["student"].queryset = Student.objects.filter(is_active=True).exclude(id__in=enrolled_ids)
    session_form = CourseSessionForm()
    session_form.fields["schedule"].queryset = Schedule.objects.filter(course=course, is_active=True).select_related("classe")
    return render(request, "attendance/course_detail.html", {
        "course": course,
        "enrollments": enrollments,
        "sessions": sessions,
        "session_form": session_form,
        "enrollment_form": enrollment_form,
    })


def course_delete(request: HttpRequest, course_id: int) -> HttpResponse:
    course = get_object_or_404(Course, id=course_id)
    if request.method == "POST":
        course.delete()
        messages.success(request, f"Cours '{course.name}' supprime.")
        return redirect("attendance:course_list")
    return render(request, "attendance/course_confirm_delete.html", {"course": course})


def enrollment_add(request: HttpRequest, course_id: int) -> HttpResponse:
    course = get_object_or_404(Course, id=course_id)
    if request.method == "POST":
        form = EnrollmentForm(request.POST)
        if form.is_valid():
            student = form.cleaned_data["student"]
            _, created = Enrollment.objects.get_or_create(student=student, course=course)
            if created:
                messages.success(request, f"{student.full_name} inscrit au cours {course.code}.")
            else:
                messages.warning(request, f"{student.full_name} est deja inscrit.")
    return redirect("attendance:course_detail", course_id=course_id)


def enrollment_remove(request: HttpRequest, enrollment_id: int) -> HttpResponse:
    enrollment = get_object_or_404(Enrollment, id=enrollment_id)
    course_id = enrollment.course_id
    if request.method == "POST":
        name = enrollment.student.full_name
        enrollment.delete()
        messages.success(request, f"{name} desincrit.")
    return redirect("attendance:course_detail", course_id=course_id)


# ─────────────────────────────────────────────────────────────────────────────
# SESSIONS
# ─────────────────────────────────────────────────────────────────────────────

def session_create(request: HttpRequest, course_id: int) -> HttpResponse:
    course = get_object_or_404(Course, id=course_id)
    if request.method == "POST":
        form = CourseSessionForm(request.POST)
        if form.is_valid():
            session_date = form.cleaned_data.get("date")
            # Verifier si jour ferie
            if session_date and JourFerie.is_ferie(session_date):
                jour = JourFerie.objects.get(date=session_date)
                messages.error(
                    request,
                    f"Impossible de creer une session le {session_date:%d/%m/%Y} : "
                    f"c'est un jour ferie ({jour.nom})."
                )
                return redirect("attendance:course_detail", course_id=course_id)

            session = form.save(commit=False)
            session.course = course
            if session.schedule:
                if not session.end_time:
                    session.end_time = session.schedule.heure_fin
                session.late_after_minutes = session.schedule.tolerance_retard_minutes
                session.minutes_avant_cours = session.schedule.minutes_avant_cours
                if not session.room and session.schedule.salle:
                    session.room = str(session.schedule.salle)
            session.save()
            messages.success(request, f"Session du {session.date:%d/%m/%Y} creee.")
        else:
            messages.error(request, f"Erreur : {form.errors.as_text()}")
    return redirect("attendance:course_detail", course_id=course_id)


def session_detail(request: HttpRequest, session_id: int) -> HttpResponse:
    session = get_object_or_404(
        CourseSession.objects.select_related("course", "schedule__classe"), id=session_id
    )
    enrolled_students = list(session.get_enrolled_students().select_related("classe"))
    records = {
        r.student_id: r
        for r in AttendanceRecord.objects.filter(course_session=session).select_related("student")
    }
    present_ids = {sid for sid, r in records.items() if r.status in ("present", "late", "excuse")}
    absent_count = max(0, len(enrolled_students) - len(present_ids))
    rows = [{"student": s, "record": records.get(s.id)} for s in enrolled_students]
    cancel_form = SessionCancelForm()
    return render(request, "attendance/session_detail.html", {
        "session": session,
        "rows": rows,
        "present_count": len(present_ids),
        "absent_count": absent_count,
        "enrolled_count": len(enrolled_students),
        "cancel_form": cancel_form,
    })


def session_open(request: HttpRequest, session_id: int) -> HttpResponse:
    session = get_object_or_404(CourseSession, id=session_id)
    if request.method == "POST":
        # Verifier jour ferie
        if JourFerie.is_ferie(session.date):
            jour = JourFerie.objects.get(date=session.date)
            messages.error(request, f"Impossible d'ouvrir : aujourd'hui est un jour ferie ({jour.nom}).")
            return redirect("attendance:session_detail", session_id=session_id)
        session.status = CourseSession.STATUS_OUVERT
        session.closed = False
        session.save(update_fields=["status", "closed"])
        messages.success(request, "Session ouverte — la reconnaissance faciale est maintenant active.")
    return redirect("attendance:session_detail", session_id=session_id)


def session_close(request: HttpRequest, session_id: int) -> HttpResponse:
    session = get_object_or_404(CourseSession, id=session_id)
    if request.method == "POST":
        count = session.generate_absents()
        session.status = CourseSession.STATUS_FERME
        session.closed = True
        session.save(update_fields=["status", "closed"])
        messages.success(request, f"Session fermee — {count} absence(s) generee(s) automatiquement.")
    return redirect("attendance:session_detail", session_id=session_id)


def session_cancel(request: HttpRequest, session_id: int) -> HttpResponse:
    session = get_object_or_404(CourseSession, id=session_id)
    if request.method == "POST":
        form = SessionCancelForm(request.POST)
        if form.is_valid():
            session.status = CourseSession.STATUS_ANNULE
            session.motif_annulation = form.cleaned_data["motif_annulation"]
            session.closed = True
            session.save(update_fields=["status", "motif_annulation", "closed"])
            messages.success(request, f"Session annulee : {session.motif_annulation}")
        else:
            messages.error(request, "Veuillez indiquer un motif d'annulation.")
    return redirect("attendance:session_detail", session_id=session_id)


def session_mark_absent(request: HttpRequest, session_id: int) -> HttpResponse:
    session = get_object_or_404(CourseSession, id=session_id)
    if request.method == "POST":
        count = session.generate_absents()
        session.status = CourseSession.STATUS_FERME
        session.closed = True
        session.save(update_fields=["status", "closed"])
        messages.success(request, f"{count} eleve(s) marque(s) absent(s). Session fermee.")
    return redirect("attendance:session_detail", session_id=session_id)


def session_delete(request: HttpRequest, session_id: int) -> HttpResponse:
    session = get_object_or_404(CourseSession, id=session_id)
    course_id = session.course_id
    if request.method == "POST":
        session.delete()
        messages.success(request, "Session supprimee.")
    return redirect("attendance:course_detail", course_id=course_id)


# ─────────────────────────────────────────────────────────────────────────────
# CAMERAS
# ─────────────────────────────────────────────────────────────────────────────

def camera_list(request: HttpRequest) -> HttpResponse:
    if request.method == "POST":
        form = CameraForm(request.POST)
        if form.is_valid():
            cam = form.save()
            _syslog(request, SystemLog.ACTION_CREATE, SystemLog.OBJ_CAMERA,
                    cam.id, cam.name, f"Source: {cam.source_display} | Zone: {cam.zone_type}")
            messages.success(request, f"Camera '{cam.name}' ajoutee.")
            return redirect("attendance:camera_list")
        messages.error(request, "Erreur dans le formulaire.")
    else:
        form = CameraForm()
    cameras = Camera.objects.select_related("salle").all()
    return render(request, "attendance/camera_list.html", {"cameras": cameras, "form": form})


def camera_edit(request: HttpRequest, camera_id: int) -> HttpResponse:
    camera = get_object_or_404(Camera, id=camera_id)
    if request.method == "POST":
        form = CameraForm(request.POST, instance=camera)
        if form.is_valid():
            form.save()
            _syslog(request, SystemLog.ACTION_UPDATE, SystemLog.OBJ_CAMERA,
                    camera.id, camera.name, f"Zone: {camera.zone_type} | Mode: {camera.detection_mode}")
            messages.success(request, "Camera mise a jour.")
            return redirect("attendance:camera_list")
        messages.error(request, "Erreur dans le formulaire.")
    else:
        form = CameraForm(instance=camera)
    return render(request, "attendance/camera_edit.html", {"form": form, "camera": camera})


def camera_delete(request: HttpRequest, camera_id: int) -> HttpResponse:
    camera = get_object_or_404(Camera, id=camera_id)
    if request.method == "POST":
        name = camera.name
        camera.delete()
        _syslog(request, SystemLog.ACTION_DELETE, SystemLog.OBJ_CAMERA, camera_id, name)
        messages.success(request, "Camera supprimee.")
        return redirect("attendance:camera_list")
    return render(request, "attendance/camera_confirm_delete.html", {"camera": camera})


def camera_live(request: HttpRequest, camera_id: int) -> HttpResponse:
    from .models import Classe
    camera = get_object_or_404(Camera, id=camera_id)
    classes = Classe.objects.filter(is_active=True).order_by("niveau", "nom")
    return render(request, "attendance/camera_live.html", {
        "camera": camera,
        "classes": classes,
    })


def camera_monitor(request: HttpRequest) -> HttpResponse:
    from .models import Classe
    cameras = Camera.objects.filter(is_active=True).order_by("name")
    classes = Classe.objects.filter(is_active=True).order_by("niveau", "nom")
    return render(request, "attendance/camera_monitor.html", {
        "cameras": cameras,
        "classes": classes,
        "cameras_json": json.dumps([
            {
                "id": c.id,
                "name": c.name,
                "location": c.location,
                "type": c.camera_type,
                "source": c.source,
                "resolution_w": c.resolution_w,
                "resolution_h": c.resolution_h,
            }
            for c in cameras
        ]),
    })


# ─────────────────────────────────────────────────────────────────────────────
# API — Frame recognition (AJAX)
# ─────────────────────────────────────────────────────────────────────────────

@csrf_exempt
def api_recognize_frame(request: HttpRequest) -> JsonResponse:
    if request.method != "POST":
        return JsonResponse({"error": "POST requis"}, status=405)

    try:
        payload = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({"error": "JSON invalide"}, status=400)

    image_b64: str = payload.get("image", "")
    session_id: str = payload.get("session_id", "")
    camera_id: str = payload.get("camera_id", "")
    classe_id: str = str(payload.get("classe_id", "") or "").strip()

    if not image_b64:
        return JsonResponse({"error": "Pas d'image"}, status=400)

    if "," in image_b64:
        image_b64 = image_b64.split(",", 1)[1]

    try:
        image_bytes = base64.b64decode(image_b64)
    except Exception:
        return JsonResponse({"error": "Image base64 invalide"}, status=400)

    session: CourseSession | None = None
    if session_id:
        session = CourseSession.objects.filter(
            id=session_id, status=CourseSession.STATUS_OUVERT
        ).first()

    camera: Camera | None = None
    if camera_id:
        camera = Camera.objects.filter(id=camera_id).first()

    allowed_student_ids = None
    selected_classe = None
    if classe_id:
        selected_classe = Classe.objects.filter(id=classe_id, is_active=True).first()
        if selected_classe:
            allowed_student_ids = frozenset(
                Student.objects.filter(classe=selected_classe, is_active=True).values_list("id", flat=True)
            )

    # ── Vérifie qu'on peut reconnaître (model LBPH OU embeddings en DB) ───────
    from .models import FaceEmbedding as _FE
    has_embeddings = _FE.objects.filter(student__is_active=True).exists()
    if not _model_ready() and not has_embeddings:
        return JsonResponse({
            "error": "Aucun modèle de reconnaissance disponible. Ajoutez des photos aux élèves pour activer la reconnaissance par empreintes faciales.",
            "results": [],
        })

    import sys as _sys
    try:
        results = recognize_from_image_bytes(
            image_bytes,
            session=session,
            allowed_student_ids=allowed_student_ids,
            camera=camera,
            source="live",
            raise_if_no_face=False,
        )
        print(f"[API] frame={len(image_bytes)}o → {len(results)} visage(s)", file=_sys.stderr)
    except RuntimeError as exc:
        return JsonResponse({"error": str(exc), "results": [], "faces_detected": 0})
    except Exception as exc:
        return JsonResponse({"error": f"Erreur interne : {exc}", "results": [], "faces_detected": 0})

    # ── Heartbeat : mettre à jour le statut de la caméra ─────────────────────
    if camera:
        camera.mark_online(fps=payload.get("fps", 0.0))

    # ── Vérifier le mode de détection ────────────────────────────────────────
    if camera and camera.detection_mode == Camera.MODE_OFF:
        return JsonResponse({"results": [], "saved": 0, "error": None, "mode": "off",
                             "info": "Camera en mode OFF — aucun traitement."})
    if camera and camera.detection_mode == Camera.MODE_MONITORING:
        return JsonResponse({"results": [], "saved": 0, "error": None, "mode": "monitoring",
                             "info": "Camera en mode MONITORING — surveillance sans reconnaissance."})

    if camera and camera.zone_type == Camera.ZONE_CHECK_OUT:
        return JsonResponse({
            "results": [],
            "saved": 0,
            "error": None,
            "mode": "monitoring",
            "info": "La sortie n'est pas prise en compte dans la presence journaliere.",
            "faces_detected": len(results),
        })

    # ── Mode présence journalière ────────────────────────────────────────────
    # Une caméra CHECK-IN enregistre toujours l'arrivée. Sans session explicite,
    # la reconnaissance live doit aussi alimenter la présence journalière, sinon
    # un mauvais paramétrage de zone donne l'impression que "rien ne se passe".
    daily_mode = (camera and camera.zone_type == Camera.ZONE_CHECK_IN) or session is None
    if daily_mode:
        day_config = SchoolDayConfig.get()
        saved = 0
        result_data = []
        for r in results:
            if r.status != "recognized" or r.student is None:
                if r.status == "unknown" and r.face_bytes:
                    log = UnknownFaceLog(source="live", camera=camera)
                    log.image.save(
                        f"unknown_live_{timezone.now().strftime('%Y%m%d_%H%M%S_%f')}.jpg",
                        ContentFile(r.face_bytes), save=True,
                    )
                attendance_status = "not_in_class" if r.status == "not_in_class" else "unknown"
                result_data.append({
                    "status": r.status,
                    "attendance_status": attendance_status,
                    "attendance_label": "Hors classe" if r.status == "not_in_class" else "Inconnu",
                    "name": r.student.full_name if r.student else "Inconnu",
                    "student_code": r.student.student_code if r.student else "",
                    "classroom": r.student.classe_display if r.student else "",
                    "confidence": r.confidence,
                    "already_marked": False,
                    "refused": False,
                    "too_early": False,
                    "bbox": [r.bbox_x_pct, r.bbox_y_pct, r.bbox_w_pct, r.bbox_h_pct],
                    "mode": "daily",
                    "debug_chi2": r.debug_chi2,
                    "debug_cos": r.debug_cos,
                    "debug_combined": round(100.0 - r.distance_lbph, 1),
                    "debug_distance": r.distance_lbph,
                    "debug_margin": r.debug_margin,
                    "debug_engine": r.debug_engine,
                    "debug_reason": r.debug_reason,
                })
                continue

            outcome = register_check_in(
                r.student,
                camera,
                config=day_config,
                source="live",
                confidence=r.confidence,
            )
            record, op = outcome.record, outcome.op

            if op == "doublon":
                attendance_status = record.status if record else DailyAttendance.STATUS_PRESENT
                result_data.append({
                    "status": "already_marked",
                    "attendance_status": attendance_status,
                    "attendance_label": dict(DailyAttendance.STATUS_CHOICES).get(attendance_status, "Deja enregistre"),
                    "name": r.student.full_name,
                    "student_code": r.student.student_code,
                    "classroom": r.student.classe_display,
                    "confidence": r.confidence,
                    "already_marked": True,
                    "refused": False,
                    "too_early": False,
                    "bbox": [r.bbox_x_pct, r.bbox_y_pct, r.bbox_w_pct, r.bbox_h_pct],
                    "mode": "daily",
                })
            elif op == "after_end":
                attendance_status = record.status if record else DailyAttendance.STATUS_ABSENT
                result_data.append({
                    "status": "refused",
                    "attendance_status": attendance_status,
                    "attendance_label": dict(DailyAttendance.STATUS_CHOICES).get(attendance_status, "Absent"),
                    "name": r.student.full_name,
                    "student_code": r.student.student_code,
                    "classroom": r.student.classe_display,
                    "confidence": r.confidence,
                    "already_marked": False,
                    "refused": True,
                    "too_early": False,
                    "heure_entree": str(record.heure_entree) if record and record.heure_entree else None,
                    "bbox": [r.bbox_x_pct, r.bbox_y_pct, r.bbox_w_pct, r.bbox_h_pct],
                    "mode": "daily",
                    "info": outcome.message or "Fin des cours atteinte: les non traces sont marques absents",
                })
            elif op == "manual_excuse":
                result_data.append({
                    "status": "already_marked",
                    "attendance_status": record.status,
                    "attendance_label": dict(DailyAttendance.STATUS_CHOICES).get(record.status, "Excuse"),
                    "name": r.student.full_name,
                    "student_code": r.student.student_code,
                    "classroom": r.student.classe_display,
                    "confidence": r.confidence,
                    "already_marked": True,
                    "refused": False,
                    "too_early": False,
                    "bbox": [r.bbox_x_pct, r.bbox_y_pct, r.bbox_w_pct, r.bbox_h_pct],
                    "mode": "daily",
                    "info": "Presence justifiee manuellement, non remplacee par la camera",
                })
            elif record is None:
                result_data.append({
                    "status": "refused",
                    "attendance_status": "refused",
                    "attendance_label": "Refuse",
                    "name": r.student.full_name,
                    "student_code": r.student.student_code,
                    "classroom": r.student.classe_display,
                    "confidence": r.confidence,
                    "already_marked": False,
                    "refused": True,
                    "too_early": False,
                    "bbox": [r.bbox_x_pct, r.bbox_y_pct, r.bbox_w_pct, r.bbox_h_pct],
                    "mode": "daily",
                    "info": op,
                })
            elif record is not None:
                if outcome.saved:
                    saved += 1
                result_data.append({
                    "status": record.status,
                    "attendance_status": record.status,
                    "attendance_label": dict(DailyAttendance.STATUS_CHOICES).get(record.status, record.status),
                    "name": r.student.full_name,
                    "student_code": r.student.student_code,
                    "classroom": r.student.classe_display,
                    "confidence": r.confidence,
                    "already_marked": False,
                    "refused": False,
                    "too_early": False,
                    "heure_entree": str(record.heure_entree) if record.heure_entree else None,
                    "bbox": [r.bbox_x_pct, r.bbox_y_pct, r.bbox_w_pct, r.bbox_h_pct],
                    "mode": "daily",
                })

        from .services import sface as _sface
        use_sface = any(r.debug_engine == "sface" for r in results)
        return JsonResponse({
            "results": result_data,
            "saved": saved,
            "error": None,
            "mode": "daily",
            "faces_detected": len(results),
            "debug_threshold": _sface.cosine_to_confidence(_sface.SFACE_AUTO_ACCEPT_THRESHOLD) if use_sface else 72,
            "debug_margin_threshold": _sface.SFACE_AMBIGUITY_MARGIN if use_sface else 0,
            "debug_engine": "sface" if use_sface else "lbp",
        })

    # ── Mode cours : n'enregistre que si une session explicite est fournie ───
    saved = 0
    if session is not None:
        saved = save_attendance_from_results(results, course_session=session, camera=camera, source="live")

    for r in results:
        if r.status == "unknown" and r.face_bytes:
            log = UnknownFaceLog(source="live", camera=camera)
            log.image.save(
                f"unknown_live_{timezone.now().strftime('%Y%m%d_%H%M%S_%f')}.jpg",
                ContentFile(r.face_bytes),
                save=True,
            )

    result_data = []
    for r in results:
        display_status = r.status
        if r.refused:
            display_status = "refused"
        elif r.too_early:
            display_status = "too_early"
        elif r.already_marked:
            display_status = "already_marked"
        result_data.append({
            "status": display_status,
            "name": r.student.full_name if r.student else "Inconnu",
            "student_code": r.student.student_code if r.student else "",
            "classroom": r.student.classe_display if r.student else "",
            "confidence": r.confidence,
            "already_marked": r.already_marked,
            "refused": r.refused,
            "too_early": r.too_early,
            "bbox": [r.bbox_x_pct, r.bbox_y_pct, r.bbox_w_pct, r.bbox_h_pct],
            # Scores de debug — affichés dans l'overlay et le panneau de diagnostic
            "debug_chi2": r.debug_chi2,
            "debug_cos": r.debug_cos,
            "debug_combined": r.confidence if r.confidence > 0 else round(100.0 - r.distance_lbph, 1),
            "debug_distance": r.distance_lbph,
            "debug_margin": r.debug_margin,
            "debug_engine": r.debug_engine,
            "debug_reason": r.debug_reason,
        })

    from .services import sface as _sface
    from .services.embedding import SIMILARITY_THRESHOLD as _THRESH
    _cfg_live = SystemConfig.get()
    use_sface = any(r.debug_engine == "sface" for r in results)
    return JsonResponse({
        "results": result_data,
        "saved": saved,
        "error": None,
        "debug_threshold": _sface.cosine_to_confidence(_sface.SFACE_AUTO_ACCEPT_THRESHOLD) if use_sface else _THRESH,
        "debug_margin_threshold": _sface.SFACE_AMBIGUITY_MARGIN if use_sface else 0,
        "debug_engine": "sface" if use_sface else "lbp",
        "debug_seuil_haute": getattr(_cfg_live, "seuil_confiance_haute", 40.0),
        "faces_detected": len(results),
    })


# ─────────────────────────────────────────────────────────────────────────────
# DIAGNOSTIC DE RECONNAISSANCE
# ─────────────────────────────────────────────────────────────────────────────

def diagnostic_view(request: HttpRequest) -> HttpResponse:
    """
    Page de diagnostic — montre les scores de similarité (chi2, cosinus, combiné)
    pour TOUS les élèves enregistrés, visage par visage.
    Permet de comprendre exactement pourquoi un visage est reconnu ou rejeté.
    """
    from .models import Camera as _Cam
    cameras = _Cam.objects.filter(is_active=True, camera_type="webcam")
    from .services.embedding import SIMILARITY_THRESHOLD
    cfg_diag = SystemConfig.get()
    return render(request, "attendance/diagnostic.html", {
        "cameras": cameras,
        "threshold": SIMILARITY_THRESHOLD,
        "seuil_haute": getattr(cfg_diag, "seuil_confiance_haute", 40.0),
    })


@require_POST
def api_diagnostic(request: HttpRequest) -> JsonResponse:
    """
    API JSON de diagnostic.
    Reçoit une image base64, retourne pour chaque visage détecté le tableau
    complet des scores (chi2, cosinus, combiné) pour chaque élève enregistré.
    N'écrit aucune présence — lecture seule.
    """
    import cv2 as _cv2
    import numpy as _np
    from .services import sface as _sface
    from .services.embedding import (
        AMBIGUITY_MARGIN as _MARGIN,
        AUTO_ACCEPT_THRESHOLD as _AUTO_THRESH,
        compute_face_embedding,
        find_all_scores as _find_all_scores,
        SIMILARITY_THRESHOLD as _THRESH,
    )
    from .services.recognition import _load_all_embeddings, _load_sface_embeddings, detect_faces, _sharpen_gray
    from .models import FaceEmbedding as _FE, Student as _St

    try:
        payload = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({"error": "JSON invalide"}, status=400)

    image_b64 = payload.get("image", "")
    if not image_b64:
        return JsonResponse({"error": "Pas d'image"}, status=400)
    if "," in image_b64:
        image_b64 = image_b64.split(",", 1)[1]
    try:
        image_bytes = base64.b64decode(image_b64)
    except Exception:
        return JsonResponse({"error": "Image base64 invalide"}, status=400)

    if not _FE.objects.filter(student__is_active=True).exists():
        return JsonResponse({
            "error": "Aucun embedding en base. Ajoutez des photos aux élèves.",
            "faces": [],
        })

    nparr = _np.frombuffer(image_bytes, _np.uint8)
    frame = _cv2.imdecode(nparr, _cv2.IMREAD_COLOR)
    if frame is None:
        return JsonResponse({"error": "Image invalide", "faces": []})

    img_h, img_w = frame.shape[:2]
    use_sface = _sface.models_available()
    gray_raw = _cv2.cvtColor(frame, _cv2.COLOR_BGR2GRAY)
    gray = _sharpen_gray(gray_raw)
    clahe = _cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
    gray = clahe.apply(gray)

    sface_detections = _sface.detect_faces(frame) if use_sface else []
    faces_boxes = [d.bbox for d in sface_detections] if use_sface else detect_faces(gray_raw, img_w, img_h)
    if not faces_boxes:
        return JsonResponse({"faces": [], "info": "Aucun visage détecté dans l'image."})

    all_embeddings = _load_sface_embeddings() if use_sface else _load_all_embeddings()
    student_map = {
        s.id: {"name": s.full_name, "code": s.student_code, "class": s.classe_display}
        for s in _St.objects.filter(is_active=True).select_related("classe")
    }

    _cfg_diag = SystemConfig.get()
    seuil_haute = getattr(_cfg_diag, "seuil_confiance_haute", 40.0)

    faces_result = []
    for idx_face, (x, y, w, h) in enumerate(faces_boxes):
        if use_sface:
            query_vec = _sface.extract_feature(frame, sface_detections[idx_face])
            all_scores = [
                {
                    "student_id": c.student_id,
                    "chi2_score": c.best_confidence,
                    "cos_score": round(c.cosine * 100.0, 1),
                    "combined": c.confidence,
                    "best_score": c.best_confidence,
                    "margin": round(c.margin, 3),
                    "samples": c.samples,
                    "_auto": _sface.is_auto_match(c),
                    "_low": c.cosine >= _sface.SFACE_MATCH_THRESHOLD and not _sface.is_auto_match(c),
                }
                for c in _sface.rank_matches(query_vec, all_embeddings)
            ]
            auto_threshold = _sface.cosine_to_confidence(_sface.SFACE_AUTO_ACCEPT_THRESHOLD)
            margin_threshold = _sface.SFACE_AMBIGUITY_MARGIN
        else:
            crop_gray = gray[y: y + h, x: x + w]
            query_vec = compute_face_embedding(crop_gray)
            all_scores = _find_all_scores(query_vec, all_embeddings)
            auto_threshold = _AUTO_THRESH
            margin_threshold = _MARGIN

        scored = []
        for s in all_scores:
            sid = s["student_id"]
            info = student_map.get(sid, {"name": f"Élève #{sid}", "code": "", "class": ""})
            combined = s["combined"]
            margin = s.get("margin", 0.0)
            distance = round(100.0 - combined, 1)
            scored.append({
                "student_id": sid,
                "name": info["name"],
                "code": info["code"],
                "class": info["class"],
                "chi2": s["chi2_score"],
                "cos": s["cos_score"],
                "combined": combined,
                "best_score": s.get("best_score", combined),
                "margin": margin,
                "samples": s.get("samples", 0),
                "distance": distance,
                "is_match": s.get("_auto", combined >= auto_threshold and margin >= margin_threshold),
                "low_conf": s.get("_low", combined >= _THRESH and (combined < auto_threshold or margin < margin_threshold)),
            })

        faces_result.append({
            "bbox": [
                round(x / img_w * 100, 2),
                round(y / img_h * 100, 2),
                round(w / img_w * 100, 2),
                round(h / img_h * 100, 2),
            ],
            "scores": scored,
        })

    return JsonResponse({
        "faces": faces_result,
        "threshold": _THRESH,
        "auto_threshold": auto_threshold if faces_result else (_sface.cosine_to_confidence(_sface.SFACE_AUTO_ACCEPT_THRESHOLD) if use_sface else _AUTO_THRESH),
        "margin_threshold": _sface.SFACE_AMBIGUITY_MARGIN if use_sface else _MARGIN,
        "engine": "opencv_sface" if use_sface else "lbp_histogram",
        "seuil_haute": seuil_haute,
    })


# ─────────────────────────────────────────────────────────────────────────────
# SCHEDULES
# ─────────────────────────────────────────────────────────────────────────────

def schedule_list(request: HttpRequest) -> HttpResponse:
    if request.method == "POST":
        form = ScheduleForm(request.POST)
        if form.is_valid():
            try:
                schedule = form.save()
                messages.success(request, f"Horaire enregistre : {schedule}.")
            except Exception as exc:
                messages.error(request, f"Erreur : {exc}")
            return redirect("attendance:schedule_list")
        for field, errors in form.errors.items():
            for e in errors:
                messages.error(request, f"{e}")
    else:
        form = ScheduleForm()
    schedules = Schedule.objects.select_related("classe", "course", "salle").order_by(
        "classe__niveau", "classe__nom", "jour_semaine", "heure_debut"
    )
    return render(request, "attendance/schedule_list.html", {
        "schedules": schedules,
        "form": form,
        "classes": Classe.objects.filter(is_active=True),
    })


def schedule_delete(request: HttpRequest, schedule_id: int) -> HttpResponse:
    schedule = get_object_or_404(Schedule, id=schedule_id)
    if request.method == "POST":
        schedule.delete()
        messages.success(request, "Horaire supprime.")
    return redirect("attendance:schedule_list")


# ─────────────────────────────────────────────────────────────────────────────
# ATTENDANCE LIST + EDITION MANUELLE
# ─────────────────────────────────────────────────────────────────────────────

def attendance_list(request: HttpRequest) -> HttpResponse:
    qs = AttendanceRecord.objects.select_related("student__classe", "course_session__course").order_by("-recognized_at")
    q = request.GET.get("q", "").strip()
    status = request.GET.get("status", "")
    date_from = request.GET.get("date_from", "")
    date_to = request.GET.get("date_to", "")
    classe_id = request.GET.get("classe", "")
    if q:
        qs = qs.filter(Q(student_name_snapshot__icontains=q) | Q(classroom_snapshot__icontains=q))
    if status:
        qs = qs.filter(status=status)
    if date_from:
        qs = qs.filter(recognized_at__date__gte=date_from)
    if date_to:
        qs = qs.filter(recognized_at__date__lte=date_to)
    if classe_id:
        qs = qs.filter(student__classe_id=classe_id)
    page_obj = Paginator(qs, 30).get_page(request.GET.get("page", 1))
    return render(request, "attendance/attendance_list.html", {
        "page_obj": page_obj, "q": q, "status_filter": status,
        "date_from": date_from, "date_to": date_to,
        "status_choices": AttendanceRecord.STATUS_CHOICES,
        "classes": Classe.objects.filter(is_active=True),
        "classe_filter": classe_id,
    })


def attendance_edit(request: HttpRequest, record_id: int) -> HttpResponse:
    """Modification manuelle d'un enregistrement de presence avec journal d'audit."""
    record = get_object_or_404(
        AttendanceRecord.objects.select_related("student__classe", "course_session__course"),
        id=record_id,
    )
    audit_logs = record.audit_logs.all()

    if request.method == "POST":
        form = AttendanceManualEditForm(request.POST)
        if form.is_valid():
            ancienne_valeur = record.status
            nouveau_statut = form.cleaned_data["nouveau_statut"]
            excuse_reason = form.cleaned_data.get("excuse_reason", "")
            excuse_notes = form.cleaned_data.get("excuse_notes", "")
            modifie_par = form.cleaned_data["modifie_par"]
            raison = form.cleaned_data.get("raison", "")

            # Appliquer les changements
            record.status = nouveau_statut
            record.excuse_reason = excuse_reason
            record.excuse_notes = excuse_notes
            record.modified_by = modifie_par
            record.save(update_fields=["status", "excuse_reason", "excuse_notes", "modified_by"])

            # Creer entree dans le journal d'audit
            AttendanceAuditLog.objects.create(
                attendance_record=record,
                modifie_par=modifie_par,
                ancienne_valeur=ancienne_valeur,
                nouvelle_valeur=nouveau_statut,
                raison=raison or excuse_notes,
            )

            messages.success(
                request,
                f"Presence de {record.student_name_snapshot} modifiee : "
                f"{dict(AttendanceRecord.STATUS_CHOICES).get(ancienne_valeur)} → "
                f"{dict(AttendanceRecord.STATUS_CHOICES).get(nouveau_statut)}."
            )
            # Retour a la session si possible
            if record.course_session_id:
                return redirect("attendance:session_detail", session_id=record.course_session_id)
            return redirect("attendance:attendance_list")
        messages.error(request, "Erreur dans le formulaire.")
    else:
        form = AttendanceManualEditForm(initial={
            "nouveau_statut": record.status,
            "excuse_reason": record.excuse_reason,
            "excuse_notes": record.excuse_notes,
        })

    return render(request, "attendance/attendance_edit.html", {
        "record": record,
        "form": form,
        "audit_logs": audit_logs,
    })


# ─────────────────────────────────────────────────────────────────────────────
# REPORTS & EXPORT
# ─────────────────────────────────────────────────────────────────────────────

def _build_report_rows(students_qs, q_date_from="", q_date_to=""):
    rows = []
    for s in students_qs:
        recs = s.daily_attendances.all()
        if q_date_from:
            recs = recs.filter(date__gte=q_date_from)
        if q_date_to:
            recs = recs.filter(date__lte=q_date_to)
        total = recs.filter(status__in=[
            DailyAttendance.STATUS_PRESENT,
            DailyAttendance.STATUS_RETARD,
            DailyAttendance.STATUS_ABSENT,
            DailyAttendance.STATUS_EXCUSE,
        ]).count()
        present = recs.filter(status=DailyAttendance.STATUS_PRESENT).count()
        absent = recs.filter(status=DailyAttendance.STATUS_ABSENT).count()
        late = recs.filter(status=DailyAttendance.STATUS_RETARD).count()
        excuse = recs.filter(status=DailyAttendance.STATUS_EXCUSE).count()
        present_like = present + late + excuse
        rate = round(present_like / total * 100, 1) if total > 0 else 0.0
        rows.append({
            "student": s, "total": total, "present": present,
            "absent": absent, "late": late, "excuse": excuse, "rate": rate,
        })
    rows.sort(key=lambda r: r["rate"])
    return rows


def report_view(request: HttpRequest) -> HttpResponse:
    q_student = request.GET.get("q_student", "").strip()
    q_faculty = request.GET.get("faculty", "")
    q_classe = request.GET.get("classe", "")
    q_date_from = request.GET.get("date_from", "")
    q_date_to = request.GET.get("date_to", "")

    students_qs = Student.objects.filter(is_active=True).select_related("classe")
    if q_student:
        students_qs = students_qs.filter(Q(full_name__icontains=q_student) | Q(student_code__icontains=q_student))
    if q_faculty:
        students_qs = students_qs.filter(faculty=q_faculty)
    if q_classe:
        students_qs = students_qs.filter(classe_id=q_classe)

    rows = _build_report_rows(students_qs, q_date_from, q_date_to)

    if request.GET.get("export") == "csv":
        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = 'attachment; filename="rapport_presences.csv"'
        response.write("\ufeff")
        writer = csv.writer(response, delimiter=";")
        writer.writerow(["Nom", "Matricule", "Faculte", "Classe", "Total", "Present", "Retard", "Excuse", "Absent", "Taux (%)"])
        for r in rows:
            writer.writerow([
                r["student"].full_name, r["student"].student_code,
                r["student"].faculty_display, r["student"].classe_display,
                r["total"], r["present"], r["late"], r["excuse"], r["absent"], r["rate"],
            ])
        return response

    return render(request, "attendance/reports.html", {
        "rows": rows,
        "classes": Classe.objects.filter(is_active=True),
        "faculty_choices": FACULTY_CHOICES,
        "q_student": q_student, "q_faculty": q_faculty,
        "q_classe": q_classe, "q_date_from": q_date_from, "q_date_to": q_date_to,
    })


def export_excel(request: HttpRequest) -> HttpResponse:
    """Export Excel (.xlsx) de toutes les presences."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        messages.error(request, "Le module openpyxl n'est pas installe. Utilisez l'export CSV a la place.")
        return redirect("attendance:reports")

    q_classe = request.GET.get("classe", "")
    q_date_from = request.GET.get("date_from", "")
    q_date_to = request.GET.get("date_to", "")

    students_qs = Student.objects.filter(is_active=True).select_related("classe")
    if q_classe:
        students_qs = students_qs.filter(classe_id=q_classe)

    rows = _build_report_rows(students_qs, q_date_from, q_date_to)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rapport Presences"

    # En-tete
    headers = ["Nom", "Matricule", "Faculte", "Classe", "Total", "Present", "Retard", "Excuse", "Absent", "Taux (%)"]
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Donnees
    for row_idx, r in enumerate(rows, 2):
        s = r["student"]
        ws.cell(row=row_idx, column=1, value=s.full_name)
        ws.cell(row=row_idx, column=2, value=s.student_code)
        ws.cell(row=row_idx, column=3, value=s.faculty_display)
        ws.cell(row=row_idx, column=4, value=s.classe_display)
        ws.cell(row=row_idx, column=5, value=r["total"])
        ws.cell(row=row_idx, column=6, value=r["present"])
        ws.cell(row=row_idx, column=7, value=r["late"])
        ws.cell(row=row_idx, column=8, value=r["excuse"])
        ws.cell(row=row_idx, column=9, value=r["absent"])
        taux_cell = ws.cell(row=row_idx, column=10, value=r["rate"])
        # Colorier en rouge si taux < 75%
        if r["rate"] < 75:
            taux_cell.font = Font(color="DC2626", bold=True)
        elif r["rate"] >= 90:
            taux_cell.font = Font(color="16A34A", bold=True)

    # Largeurs auto
    col_widths = [30, 15, 20, 25, 8, 10, 10, 10, 10, 10]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    today = timezone.localdate()
    fname = f"rapport_presences_{today}.xlsx"
    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{fname}"'
    return response


# ─────────────────────────────────────────────────────────────────────────────
# UNKNOWN FACES
# ─────────────────────────────────────────────────────────────────────────────

def unknown_faces_view(request: HttpRequest) -> HttpResponse:
    page_obj = Paginator(UnknownFaceLog.objects.all(), 24).get_page(request.GET.get("page", 1))
    return render(request, "attendance/unknown_faces.html", {"page_obj": page_obj})


def unknown_face_delete(request: HttpRequest, log_id: int) -> HttpResponse:
    log = get_object_or_404(UnknownFaceLog, id=log_id)
    if request.method == "POST":
        log.delete()
    return redirect("attendance:unknown_faces")


# ─────────────────────────────────────────────────────────────────────────────
# TRAINING
# ─────────────────────────────────────────────────────────────────────────────

def train_model_view(request: HttpRequest) -> HttpResponse:
    if request.method != "POST":
        return redirect("attendance:dashboard")
    t_start = _time_module.monotonic()
    record = TrainingHistory(
        triggered_by=request.user.username if request.user.is_authenticated else "system",
    )
    try:
        summary = train_model()
    except Exception as exc:
        duration = round(_time_module.monotonic() - t_start, 2)
        record.success = False
        record.error_message = str(exc)[:1000]
        record.duration_seconds = duration
        record.save()
        _syslog(request, SystemLog.ACTION_TRAIN, SystemLog.OBJ_MODEL,
                details=f"ECHEC entrainement: {exc}", success=False)
        messages.error(request, f"Erreur entrainement : {exc}")
    else:
        duration = round(_time_module.monotonic() - t_start, 2)
        record.success = True
        record.nb_students = summary.students
        record.nb_photos = summary.faces
        record.nb_skipped_blurry = summary.skipped_blurry
        record.duration_seconds = duration
        record.save()
        _syslog(request, SystemLog.ACTION_TRAIN, SystemLog.OBJ_MODEL,
                object_repr="Modele LBPH",
                details=f"OK — {summary.students} etud. | {summary.faces} photos | "
                        f"{summary.skipped_blurry} floues ignorees | {duration:.1f}s")
        messages.success(
            request,
            f"Modele entraine : {summary.students} eleve(s), {summary.faces} visage(s) utiles"
            f"{' | ' + str(summary.skipped_blurry) + ' photo(s) floue(s) ignoree(s)' if summary.skipped_blurry else ''}.",
        )
    return redirect("attendance:dashboard")


# ─────────────────────────────────────────────────────────────────────────────
# ENTRAÎNEMENT ASYNCHRONE (AJAX)
# ─────────────────────────────────────────────────────────────────────────────

def api_train_status(request: HttpRequest) -> JsonResponse:
    """Retourne l'état courant de l'entraînement LBPH (JSON)."""
    from .services.train_state import get_state
    from .services.paths import MODEL_FILE
    state = get_state()
    state["model_ready"] = MODEL_FILE.exists()
    state["untrained_photos"] = TrainingPhoto.objects.filter(trained_at__isnull=True).count()
    return JsonResponse(state)


def api_retrain_async(request: HttpRequest) -> JsonResponse:
    """Déclenche un entraînement LBPH asynchrone et retourne l'état immédiatement."""
    if request.method != "POST":
        return JsonResponse({"error": "POST requis"}, status=405)
    from .services.train_state import start_training_async, get_state
    started = start_training_async("manual")
    state = get_state()
    state["started"] = started
    return JsonResponse(state)


# ─────────────────────────────────────────────────────────────────────────────
# REBUILD ALL EMBEDDINGS
# ─────────────────────────────────────────────────────────────────────────────

def rebuild_all_embeddings_view(request: HttpRequest) -> HttpResponse:
    """Supprime et recrée tous les embeddings à partir des photos existantes."""
    if request.method != "POST":
        return redirect("attendance:dashboard")

    t_start = _time_module.monotonic()
    try:
        summary = train_model()
    except Exception as exc:
        duration = round(_time_module.monotonic() - t_start, 2)
        _syslog(request, SystemLog.ACTION_TRAIN, SystemLog.OBJ_MODEL,
                details=f"ECHEC recalcul embeddings — {exc} | {duration:.1f}s", success=False)
        messages.error(request, f"Recalcul impossible : {exc}")
        return redirect("attendance:dashboard")

    duration = round(_time_module.monotonic() - t_start, 2)
    failed = max(0, summary.photos - summary.faces)
    _syslog(request, SystemLog.ACTION_TRAIN, SystemLog.OBJ_MODEL,
            details=f"Recalcul embeddings — {summary.faces}/{summary.photos} OK | {failed} echecs | {duration:.1f}s")
    messages.success(
        request,
        f"Embeddings SFace recalcules : {summary.faces}/{summary.photos} visage(s), "
        f"{summary.students} eleve(s) — {duration:.1f}s.",
    )
    return redirect("attendance:dashboard")


# ─────────────────────────────────────────────────────────────────────────────
# RECOGNITION — photo upload
# ─────────────────────────────────────────────────────────────────────────────

def recognize_upload_view(request: HttpRequest) -> HttpResponse:
    sessions = CourseSession.objects.filter(
        status=CourseSession.STATUS_OUVERT
    ).select_related("course", "schedule__classe").order_by("-date", "-start_time")[:30]
    if request.method == "POST":
        if "image" not in request.FILES:
            messages.error(request, "Veuillez selectionner une image.")
            return render(request, "attendance/recognize_upload.html", {"results": None, "sessions": sessions})

        image_data = request.FILES["image"].read()
        session_id = request.POST.get("session_id", "").strip()
        course_session: CourseSession | None = None
        if session_id:
            course_session = CourseSession.objects.filter(
                id=session_id, status=CourseSession.STATUS_OUVERT
            ).first()

        # Filtrage par classe (point 12 : seulement les eleves de la classe active)
        allowed_ids: frozenset[int] | None = None
        if course_session and SystemConfig.get().filtrer_par_classe:
            sched = course_session.schedule
            if sched and sched.classe_id:
                allowed_ids = frozenset(
                    Student.objects.filter(classe_id=sched.classe_id, is_active=True)
                    .values_list("id", flat=True)
                )
        try:
            results = recognize_from_image_bytes(
                image_data, session=course_session,
                allowed_student_ids=allowed_ids, source="photo",
            )
        except Exception as exc:
            messages.error(request, str(exc))
            return render(request, "attendance/recognize_upload.html", {"results": None, "sessions": sessions})

        saved = save_attendance_from_results(results, course_session=course_session, source="photo")

        for r in results:
            if r.status == "unknown" and r.face_bytes:
                log = UnknownFaceLog(source="photo")
                log.image.save(
                    f"unknown_{timezone.now().strftime('%Y%m%d_%H%M%S_%f')}.jpg",
                    ContentFile(r.face_bytes), save=True,
                )

        return render(request, "attendance/recognize_upload.html", {
            "results": results, "saved": saved,
            "sessions": sessions, "selected_session": course_session,
        })
    return render(request, "attendance/recognize_upload.html", {"results": None, "sessions": sessions})


def recognize_view(request: HttpRequest) -> HttpResponse:
    return redirect("attendance:recognize_upload")


# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURATION SYSTEME
# ─────────────────────────────────────────────────────────────────────────────

def system_config_view(request: HttpRequest) -> HttpResponse:
    """Page de configuration globale du systeme (singleton SystemConfig)."""
    config = SystemConfig.get()
    if request.method == "POST":
        form = SystemConfigForm(request.POST, instance=config)
        if form.is_valid():
            form.save()
            cfg = SystemConfig.get()
            _syslog(request, SystemLog.ACTION_CONFIG, SystemLog.OBJ_CONFIG, 1,
                    "Configuration systeme",
                    f"seuil_haute={cfg.seuil_confiance_haute} | seuil_lbph={cfg.seuil_distance_lbph} | "
                    f"retard={cfg.retard_minutes}min | cooldown={cfg.cooldown_detection_minutes}min")
            messages.success(request, "Configuration enregistree avec succes.")
            return redirect("attendance:system_config")
        messages.error(request, "Erreur dans le formulaire.")
    else:
        form = SystemConfigForm(instance=config)
    return render(request, "attendance/system_config.html", {
        "form": form,
        "config": config,
    })


# ─────────────────────────────────────────────────────────────────────────────
# JOURNAL D'EVENEMENTS BRUTS DE DETECTION
# ─────────────────────────────────────────────────────────────────────────────

def detection_events_view(request: HttpRequest) -> HttpResponse:
    """Journal des evenements bruts du pipeline de reconnaissance (audit technique)."""
    qs = FaceDetectionEvent.objects.select_related(
        "student__classe", "course_session__course", "camera"
    )

    etape = request.GET.get("etape", "")
    student_q = request.GET.get("q", "").strip()
    session_id = request.GET.get("session", "")

    if etape:
        qs = qs.filter(etape=etape)
    if student_q:
        qs = qs.filter(student__full_name__icontains=student_q)
    if session_id:
        qs = qs.filter(course_session_id=session_id)

    total = qs.count()

    # Compteurs par etape
    from django.db.models import Count as _Count
    etape_counts = {
        e["etape"]: e["n"]
        for e in FaceDetectionEvent.objects.values("etape").annotate(n=_Count("id"))
    }
    etape_stats = [
        {"code": code, "label": label, "count": etape_counts.get(code, 0)}
        for code, label in FaceDetectionEvent.ETAPE_CHOICES
    ]

    page_obj = Paginator(qs, 50).get_page(request.GET.get("page", 1))

    return render(request, "attendance/detection_events.html", {
        "page_obj": page_obj,
        "etape": etape,
        "student_q": student_q,
        "session_id": session_id,
        "etape_choices": FaceDetectionEvent.ETAPE_CHOICES,
        "etape_stats": etape_stats,
        "total": total,
        "etape_counts": etape_counts,
    })


# ─────────────────────────────────────────────────────────────────────────────
# STATISTIQUES PAR CLASSE
# ─────────────────────────────────────────────────────────────────────────────

def stats_classe_view(request: HttpRequest) -> HttpResponse:
    """Statistiques de presence journaliere agregees par classe."""
    date_str = request.GET.get("date", timezone.localdate().isoformat())
    try:
        target_date = datetime.date.fromisoformat(date_str)
    except ValueError:
        target_date = timezone.localdate()

    day_config = SchoolDayConfig.get()
    day_block = _daily_day_block_reason(target_date, day_config)
    absence_block = _daily_absence_generation_block_reason(
        target_date,
        timezone.localtime().time().replace(tzinfo=None, microsecond=0),
        day_config,
    )
    missing_counts_as_absent = not absence_block
    classes = Classe.objects.filter(is_active=True).prefetch_related("students")

    rows = []
    for classe in classes:
        student_ids = list(
            Student.objects.filter(classe=classe, is_active=True).values_list("id", flat=True)
        )
        nb_students = len(student_ids)

        records = DailyAttendance.objects.filter(student_id__in=student_ids, date=target_date)
        nb_present = records.filter(status=DailyAttendance.STATUS_PRESENT).count()
        nb_retard = records.filter(status=DailyAttendance.STATUS_RETARD).count()
        nb_excuse = records.filter(status=DailyAttendance.STATUS_EXCUSE).count()
        explicit_absent = records.filter(status=DailyAttendance.STATUS_ABSENT).count()
        nb_registered = records.count()
        nb_missing = max(0, nb_students - nb_registered)
        nb_absent = explicit_absent + nb_missing if missing_counts_as_absent else explicit_absent
        nb_total = nb_students
        nb_ok = nb_present + nb_retard + nb_excuse
        rate = round(nb_ok / nb_total * 100, 1) if nb_total > 0 else 0.0

        nb_sessions = CourseSession.objects.filter(
            schedule__classe=classe,
            status__in=[CourseSession.STATUS_FERME, CourseSession.STATUS_OUVERT],
        ).count()

        # Top 3 absences
        top_absents = (
            Student.objects.filter(id__in=student_ids, is_active=True)
            .annotate(nb_abs=Count("daily_attendances", filter=Q(daily_attendances__status=DailyAttendance.STATUS_ABSENT)))
            .filter(nb_abs__gt=0)
            .order_by("-nb_abs")[:3]
        )

        rows.append({
            "classe": classe,
            "nb_students": nb_students,
            "nb_sessions": nb_sessions,
            "nb_present": nb_present,
            "nb_retard": nb_retard,
            "nb_excuse": nb_excuse,
            "nb_absent": nb_absent,
            "nb_missing": nb_missing,
            "nb_total": nb_total,
            "rate": rate,
            "top_absents": top_absents,
        })

    # Trier par taux croissant (les plus mauvaises en premier)
    rows.sort(key=lambda r: r["rate"])

    return render(request, "attendance/stats_classe.html", {
        "rows": rows,
        "target_date": target_date,
        "day_block": day_block,
        "absence_block": absence_block,
        "missing_counts_as_absent": missing_counts_as_absent,
        "prev_date": (target_date - datetime.timedelta(days=1)).isoformat(),
        "next_date": (target_date + datetime.timedelta(days=1)).isoformat(),
        "total_classes": len(rows),
        "total_students": sum(r["nb_students"] for r in rows),
        "global_rate": round(
            sum(r["nb_present"] + r["nb_retard"] + r["nb_excuse"] for r in rows)
            / max(1, sum(r["nb_students"] for r in rows)) * 100,
            1,
        ),
    })


# ─────────────────────────────────────────────────────────────────────────────
# HELPERS — PRESENCE JOURNALIERE (école secondaire)
# ─────────────────────────────────────────────────────────────────────────────

def _daily_day_block_reason(date, config: "SchoolDayConfig") -> str:
    """Retourne une raison si la journée n'accepte pas les présences automatiques."""
    return day_block_reason(date, config)


def _daily_time_block_reason(now_time, config: "SchoolDayConfig") -> str:
    """Retourne une raison si l'heure courante est hors plage portail."""
    return time_block_reason(now_time, config)


def _daily_absence_generation_block_reason(date, now_time, config: "SchoolDayConfig") -> str:
    """Empêche de marquer absent avant la fin officielle des cours."""
    return absence_generation_block_reason(date, now_time, config)


def _daily_arrival_status(now_time, config: "SchoolDayConfig") -> str:
    return arrival_status(now_time, config)


def _daily_checkin(student: "Student", camera: "Camera", config: "SchoolDayConfig"):
    """Enregistre l'arrivée d'un élève. Retourne (record, op)."""
    outcome = register_check_in(
        student,
        camera,
        config=config,
        source="live",
        log_event=False,
    )
    return outcome.record, outcome.op


def daily_attendance_list(request: HttpRequest) -> HttpResponse:
    """Liste des présences journalières avec filtres."""
    date_str = request.GET.get("date", timezone.localdate().isoformat())
    try:
        target_date = datetime.date.fromisoformat(date_str)
    except ValueError:
        target_date = timezone.localdate()

    q = request.GET.get("q", "").strip()
    status_filter = request.GET.get("status", "")
    classe_filter = request.GET.get("classe", "")
    day_config = SchoolDayConfig.get()
    day_block_reason = _daily_day_block_reason(target_date, day_config)

    qs = DailyAttendance.objects.filter(date=target_date).select_related(
        "student__classe", "camera_entree"
    )
    if q:
        qs = qs.filter(
            Q(student__full_name__icontains=q) | Q(student__student_code__icontains=q)
        )
    if status_filter:
        qs = qs.filter(status=status_filter)
    if classe_filter:
        qs = qs.filter(student__classe_id=classe_filter)

    qs = qs.order_by("student__full_name")

    # Étudiants actifs sans enregistrement ce jour-là
    all_active = Student.objects.filter(is_active=True)
    if classe_filter:
        all_active = all_active.filter(classe_id=classe_filter)
    present_ids = set(
        DailyAttendance.objects.filter(date=target_date).values_list("student_id", flat=True)
    )
    absent_students = Student.objects.none() if day_block_reason else all_active.exclude(id__in=present_ids).select_related("classe").order_by("full_name")
    if q:
        absent_students = absent_students.filter(
            Q(full_name__icontains=q) | Q(student_code__icontains=q)
        )

    page_obj = Paginator(list(qs) + list(absent_students), 50).get_page(request.GET.get("page", 1))

    return render(request, "attendance/daily_attendance_list.html", {
        "target_date": target_date,
        "records": qs,
        "absent_students": absent_students,
        "total_present": qs.count(),
        "total_absent": absent_students.count(),
        "q": q,
        "status_filter": status_filter,
        "classe_filter": classe_filter,
        "classes": Classe.objects.filter(is_active=True).order_by("nom"),
        "status_choices": DailyAttendance.STATUS_CHOICES,
        "day_block_reason": day_block_reason,
        "prev_date": (target_date - datetime.timedelta(days=1)).isoformat(),
        "next_date": (target_date + datetime.timedelta(days=1)).isoformat(),
    })


def daily_attendance_edit(request: HttpRequest, record_id: int) -> HttpResponse:
    """Modification manuelle d'une présence journalière."""
    record = get_object_or_404(DailyAttendance.objects.select_related("student__classe"), id=record_id)

    if request.method == "POST":
        form = DailyAttendanceEditForm(request.POST)
        if form.is_valid():
            cd = form.cleaned_data
            record.status = cd["nouveau_statut"]
            if cd.get("heure_entree"):
                record.heure_entree = cd["heure_entree"]
            record.excuse_reason = cd.get("excuse_reason", "")
            record.excuse_notes = cd.get("excuse_notes", "")
            record.modified_by = cd.get("modifie_par", "Admin")
            record.save()
            messages.success(request, f"Presence de {record.student.full_name} mise a jour ({record.get_status_display()}).")
            return redirect("attendance:daily_attendance_list")
        messages.error(request, "Erreur dans le formulaire.")
    else:
        form = DailyAttendanceEditForm(initial={
            "nouveau_statut": record.status,
            "heure_entree": record.heure_entree,
            "excuse_reason": record.excuse_reason,
            "excuse_notes": record.excuse_notes,
            "modifie_par": record.modified_by or "Admin",
        })

    return render(request, "attendance/daily_attendance_edit.html", {
        "record": record,
        "form": form,
        "student": record.student,
    })


def daily_generate_absents(request: HttpRequest) -> HttpResponse:
    """Génère les enregistrements ABSENT pour les élèves sans présence aujourd'hui."""
    if request.method != "POST":
        return redirect("attendance:daily_attendance_list")

    today = timezone.localdate()
    day_config = SchoolDayConfig.get()
    outcome = generate_absences_for_date(
        today,
        config=day_config,
        now=timezone.now(),
        modified_by="Systeme (generation auto)",
    )
    if outcome.blocked:
        messages.error(request, f"Generation des absences impossible : {outcome.message}.")
        return redirect("attendance:daily_attendance_list")

    messages.success(request, f"{outcome.count} absence(s) generee(s) automatiquement pour le {today:%d/%m/%Y}.")
    return redirect("attendance:daily_attendance_list")


def school_day_config_view(request: HttpRequest) -> HttpResponse:
    """Configuration de la journée scolaire (école secondaire)."""
    config = SchoolDayConfig.get()
    if request.method == "POST":
        form = SchoolDayConfigForm(request.POST, instance=config)
        if form.is_valid():
            form.save()
            messages.success(request, "Configuration journaliere sauvegardee.")
            return redirect("attendance:school_day_config")
        messages.error(request, "Erreur dans le formulaire.")
    else:
        form = SchoolDayConfigForm(instance=config)

    return render(request, "attendance/school_day_config.html", {
        "form": form,
        "config": config,
    })


# ─────────────────────────────────────────────────────────────────────────────
# CONTRÔLE DES CAMÉRAS — Panneau admin (École secondaire)
# ─────────────────────────────────────────────────────────────────────────────

def camera_control_panel(request: HttpRequest) -> HttpResponse:
    """Panneau de contrôle centralisé pour toutes les caméras."""
    cameras = Camera.objects.select_related("salle").order_by("zone_type", "name")
    day_config = SchoolDayConfig.get()

    # Marquer hors-ligne les caméras sans heartbeat depuis > 2 minutes
    cutoff = timezone.now() - datetime.timedelta(minutes=2)
    Camera.objects.filter(is_online=True).filter(
        last_seen__lt=cutoff
    ).update(is_online=False)
    cameras = Camera.objects.select_related("salle").order_by("zone_type", "name")

    nb_online = cameras.filter(is_online=True).count()
    nb_recognition = cameras.filter(detection_mode=Camera.MODE_RECOGNITION, is_active=True).count()
    nb_monitoring = cameras.filter(detection_mode=Camera.MODE_MONITORING, is_active=True).count()
    nb_off = cameras.filter(detection_mode=Camera.MODE_OFF).count() + cameras.filter(is_active=False).count()

    return render(request, "attendance/camera_control_panel.html", {
        "cameras": cameras,
        "day_config": day_config,
        "nb_online": nb_online,
        "nb_recognition": nb_recognition,
        "nb_monitoring": nb_monitoring,
        "nb_off": nb_off,
        "mode_choices": Camera.MODE_CHOICES,
    })


@require_POST
def camera_set_mode(request: HttpRequest, camera_id: int) -> JsonResponse:
    """Change le mode de détection d'une caméra (AJAX ou formulaire POST)."""
    camera = get_object_or_404(Camera, id=camera_id)
    mode = request.POST.get("mode", "")
    valid_modes = [Camera.MODE_RECOGNITION, Camera.MODE_MONITORING, Camera.MODE_OFF]
    if mode not in valid_modes:
        return JsonResponse({"error": f"Mode invalide : {mode}"}, status=400)
    old_mode = camera.detection_mode
    camera.detection_mode = mode
    camera.save(update_fields=["detection_mode"])
    _syslog(request, SystemLog.ACTION_CAMERA, SystemLog.OBJ_CAMERA,
            camera.id, camera.name,
            f"Mode change: {old_mode} → {mode}")
    return JsonResponse({
        "ok": True,
        "camera_id": camera.id,
        "mode": mode,
        "mode_label": camera.get_detection_mode_display(),
        "status_label": camera.status_label,
        "status_badge": camera.status_badge,
    })


@csrf_exempt
def api_camera_heartbeat(request: HttpRequest, camera_id: int) -> JsonResponse:
    """
    Endpoint appelé par les services caméra externes pour signaler leur activité.
    POST { "fps": 12.5, "error": "" }
    """
    if request.method != "POST":
        return JsonResponse({"error": "POST requis"}, status=405)

    camera = get_object_or_404(Camera, id=camera_id)

    try:
        payload = json.loads(request.body) if request.body else {}
    except json.JSONDecodeError:
        payload = {}

    error_msg = payload.get("error", "")
    fps = float(payload.get("fps", 0.0))

    if error_msg:
        camera.mark_offline(error_msg)
        return JsonResponse({"ok": True, "status": "offline", "error": error_msg})

    camera.mark_online(fps=fps)
    return JsonResponse({
        "ok": True,
        "status": "online",
        "detection_mode": camera.detection_mode,
        "zone_type": camera.zone_type,
        "fps": fps,
    })


@csrf_exempt
def api_cameras_status(request: HttpRequest) -> JsonResponse:
    """Retourne l'état de toutes les caméras actives (polling dashboard)."""
    # Marquer hors-ligne les caméras sans heartbeat depuis > 2 minutes
    cutoff = timezone.now() - datetime.timedelta(minutes=2)
    Camera.objects.filter(is_online=True, last_seen__lt=cutoff).update(is_online=False)

    cameras = Camera.objects.filter(is_active=True).values(
        "id", "name", "location", "zone_type", "detection_mode",
        "is_online", "last_seen", "fps_estimate", "frames_processed",
        "error_count", "last_error",
    )
    data = []
    for c in cameras:
        last_seen_str = c["last_seen"].strftime("%H:%M:%S") if c["last_seen"] else None
        data.append({**c, "last_seen": last_seen_str})
    return JsonResponse({"cameras": data, "count": len(data)})


# ─────────────────────────────────────────────────────────────────────────────
# FILE DE REVUE — Reconnaissances à faible confiance
# ─────────────────────────────────────────────────────────────────────────────

def review_queue_list(request: HttpRequest) -> HttpResponse:
    """Liste des tickets de revue reconnaissance."""
    status_filter = request.GET.get("status", "pending")
    q = request.GET.get("q", "").strip()

    qs = RecognitionReviewQueue.objects.select_related(
        "student_proposed__classe", "second_candidate", "camera"
    )
    if status_filter:
        qs = qs.filter(status=status_filter)
    if q:
        qs = qs.filter(
            Q(student_proposed__full_name__icontains=q)
            | Q(student_proposed__student_code__icontains=q)
        )

    nb_pending = RecognitionReviewQueue.objects.filter(status=RecognitionReviewQueue.STATUS_PENDING).count()
    nb_validated = RecognitionReviewQueue.objects.filter(status=RecognitionReviewQueue.STATUS_VALIDATED).count()
    nb_rejected = RecognitionReviewQueue.objects.filter(status=RecognitionReviewQueue.STATUS_REJECTED).count()

    page_obj = Paginator(qs, 20).get_page(request.GET.get("page", 1))

    return render(request, "attendance/review_queue.html", {
        "tickets": page_obj,
        "page_obj": page_obj,
        "status_filter": status_filter,
        "q": q,
        "nb_pending": nb_pending,
        "nb_validated": nb_validated,
        "nb_rejected": nb_rejected,
    })


def review_queue_detail(request: HttpRequest, ticket_id: int) -> HttpResponse:
    """Détail et validation manuelle d'un ticket de revue."""
    ticket = get_object_or_404(
        RecognitionReviewQueue.objects.select_related(
            "student_proposed__classe", "second_candidate", "camera", "course_session"
        ),
        id=ticket_id,
    )

    if request.method == "POST" and ticket.status == RecognitionReviewQueue.STATUS_PENDING:
        form = ReviewQueueValidateForm(request.POST)
        action_reject = request.POST.get("action_reject")  # bouton "Rejeter"

        reviewed_by = request.POST.get("reviewed_by", "Admin").strip() or "Admin"
        review_notes = request.POST.get("review_notes", "").strip()

        if action_reject:
            # Rejet
            ticket.status = RecognitionReviewQueue.STATUS_REJECTED
            ticket.reviewed_by = reviewed_by
            ticket.reviewed_at = timezone.now()
            ticket.review_notes = review_notes
            ticket.save()
            messages.success(request, f"Ticket #{ticket.id} rejete. Aucune presence enregistree.")
            return redirect("attendance:review_queue_list")

        if form.is_valid():
            action = form.cleaned_data["action"]
            if action == ReviewQueueValidateForm.ACTION_VALIDATE and ticket.student_proposed:
                from .services.recognition import save_attendance_from_results, FaceResult
                from .services.embedding import extract_embedding_from_image, vector_to_bytes
                from .models import FaceEmbedding
                import io

                student = ticket.student_proposed
                reviewed_by = form.cleaned_data["reviewed_by"]
                review_notes = form.cleaned_data.get("review_notes", "")
                now = timezone.now()

                # ── 1. Enregistrer la présence du ticket courant ────────────────
                synthetic = FaceResult(
                    status="recognized",
                    student=student,
                    confidence=ticket.confidence_proposed,
                    distance_lbph=ticket.distance_lbph,
                )
                saved = save_attendance_from_results(
                    [synthetic],
                    course_session=ticket.course_session,
                    camera=ticket.camera,
                    source="review_validated",
                )
                ticket.status = RecognitionReviewQueue.STATUS_VALIDATED
                ticket.reviewed_by = reviewed_by
                ticket.reviewed_at = now
                ticket.review_notes = review_notes
                ticket.save()

                # ── 2. Auto-apprentissage : sauvegarder le visage comme photo d'entraînement ──
                embed_added = 0
                if ticket.face_image:
                    try:
                        ticket.face_image.open("rb")
                        image_bytes = ticket.face_image.read()
                        ticket.face_image.close()
                        photo = TrainingPhoto.objects.create(
                            student=student,
                            image=ticket.face_image,
                            angle_tag="",
                        )
                        embedding, blur_s, quality = extract_embedding_from_image(image_bytes)
                        if embedding is not None:
                            FaceEmbedding.objects.create(
                                student=student,
                                photo=photo,
                                vector=vector_to_bytes(embedding),
                                score_qualite=blur_s,
                            )
                            embed_added = 1
                    except Exception:
                        pass

                # ── 3. Validation en lot : tous les autres tickets "pending" du même élève ──
                other_pending = RecognitionReviewQueue.objects.filter(
                    student_proposed=student,
                    status=RecognitionReviewQueue.STATUS_PENDING,
                ).exclude(pk=ticket.pk)

                bulk_saved = 0
                bulk_count = other_pending.count()
                for other in other_pending:
                    other_synthetic = FaceResult(
                        status="recognized",
                        student=student,
                        confidence=other.confidence_proposed,
                        distance_lbph=other.distance_lbph,
                    )
                    s = save_attendance_from_results(
                        [other_synthetic],
                        course_session=other.course_session,
                        camera=other.camera,
                        source="review_bulk_validated",
                    )
                    bulk_saved += s
                    other.status = RecognitionReviewQueue.STATUS_VALIDATED
                    other.reviewed_by = reviewed_by
                    other.reviewed_at = now
                    other.review_notes = f"Validation en lot depuis ticket #{ticket.id}"
                    other.save()

                # ── 4. Message de confirmation ──────────────────────────────────
                msg_parts = []
                if saved:
                    msg_parts.append(f"Présence de {student.full_name} enregistrée.")
                else:
                    msg_parts.append(f"Ticket validé (présence déjà marquée ou hors-horaire).")
                if bulk_count:
                    msg_parts.append(f"{bulk_count} autre(s) ticket(s) du même élève validé(s) automatiquement ({bulk_saved} présence(s) enregistrée(s)).")
                if embed_added:
                    msg_parts.append("Visage ajouté aux photos d'entraînement — reconnaissance améliorée.")
                messages.success(request, " ".join(msg_parts))

            else:
                ticket.status = RecognitionReviewQueue.STATUS_REJECTED
                ticket.reviewed_by = form.cleaned_data["reviewed_by"]
                ticket.reviewed_at = timezone.now()
                ticket.review_notes = form.cleaned_data.get("review_notes", "")
                ticket.save()
                messages.success(request, f"Ticket #{ticket.id} rejeté.")
            return redirect("attendance:review_queue_list")
    else:
        form = ReviewQueueValidateForm(initial={"reviewed_by": request.user.username if request.user.is_authenticated else "Admin"})

    return render(request, "attendance/review_queue_detail.html", {
        "ticket": ticket,
        "form": form,
    })


# ─────────────────────────────────────────────────────────────────────────────
# GESTION DES UTILISATEURS
# ─────────────────────────────────────────────────────────────────────────────

AuthUser = get_user_model()


def _require_admin(request: HttpRequest) -> HttpResponse | None:
    """Retourne une 403 si l'utilisateur n'est pas staff. None = autorisé."""
    if not request.user.is_authenticated or not (request.user.is_staff or request.user.is_superuser):
        return HttpResponse("Accès refusé. Vous devez être administrateur.", status=403)
    return None


def user_list(request: HttpRequest) -> HttpResponse:
    guard = _require_admin(request)
    if guard:
        return guard
    users = AuthUser.objects.all().order_by("-is_superuser", "-is_staff", "username")
    nb_active = users.filter(is_active=True).count()
    nb_staff = users.filter(is_staff=True).count()
    return render(request, "attendance/user_list.html", {
        "users": users,
        "nb_active": nb_active,
        "nb_staff": nb_staff,
    })


def user_create(request: HttpRequest) -> HttpResponse:
    guard = _require_admin(request)
    if guard:
        return guard

    if request.method == "POST":
        form = UserCreateForm(request.POST)
        if form.is_valid():
            user = form.save()
            role = "super-admin" if user.is_superuser else ("staff" if user.is_staff else "utilisateur")
            _syslog(request, SystemLog.ACTION_CREATE, SystemLog.OBJ_USER,
                    user.id, user.username,
                    f"Role: {role} | Email: {user.email or 'N/A'}")
            messages.success(request, f"Utilisateur «\u00a0{user.username}\u00a0» créé avec succès.")
            return redirect("attendance:user_list")
    else:
        form = UserCreateForm()

    return render(request, "attendance/user_form.html", {
        "form": form,
        "is_create": True,
    })


def user_edit(request: HttpRequest, user_id: int) -> HttpResponse:
    guard = _require_admin(request)
    if guard:
        return guard

    target = get_object_or_404(AuthUser, pk=user_id)

    # Empêche un non-superuser de modifier un superuser
    if target.is_superuser and not request.user.is_superuser:
        messages.error(request, "Seul un super-administrateur peut modifier un autre super-administrateur.")
        return redirect("attendance:user_list")

    if request.method == "POST":
        form = UserEditForm(request.POST, instance=target)
        if form.is_valid():
            form.save()
            _syslog(request, SystemLog.ACTION_UPDATE, SystemLog.OBJ_USER,
                    target.id, target.username,
                    f"is_staff={target.is_staff} | is_superuser={target.is_superuser} | is_active={target.is_active}")
            messages.success(request, f"Compte «\u00a0{target.username}\u00a0» mis à jour.")
            return redirect("attendance:user_list")
    else:
        form = UserEditForm(instance=target)

    return render(request, "attendance/user_form.html", {
        "form": form,
        "target_user": target,
        "is_create": False,
    })


@require_POST
def user_toggle_active(request: HttpRequest, user_id: int) -> HttpResponse:
    guard = _require_admin(request)
    if guard:
        return guard

    target = get_object_or_404(AuthUser, pk=user_id)

    if target == request.user:
        messages.error(request, "Vous ne pouvez pas désactiver votre propre compte.")
        return redirect("attendance:user_list")

    target.is_active = not target.is_active
    target.save(update_fields=["is_active"])
    action_str = "réactivé" if target.is_active else "désactivé"
    _syslog(request, SystemLog.ACTION_UPDATE, SystemLog.OBJ_USER,
            target.id, target.username,
            f"Compte {action_str} (is_active={target.is_active})")
    messages.success(request, f"Compte «\u00a0{target.username}\u00a0» {action_str}.")
    return redirect("attendance:user_list")


def user_reset_password(request: HttpRequest, user_id: int) -> HttpResponse:
    guard = _require_admin(request)
    if guard:
        return guard

    target = get_object_or_404(AuthUser, pk=user_id)

    if target.is_superuser and not request.user.is_superuser:
        messages.error(request, "Seul un super-administrateur peut changer le mot de passe d'un autre super-administrateur.")
        return redirect("attendance:user_list")

    if request.method == "POST":
        form = UserResetPasswordForm(request.POST)
        if form.is_valid():
            target.set_password(form.cleaned_data["password1"])
            target.save(update_fields=["password"])
            _syslog(request, SystemLog.ACTION_UPDATE, SystemLog.OBJ_USER,
                    target.id, target.username,
                    "Mot de passe reinitialise par un administrateur")
            messages.success(request, f"Mot de passe de «\u00a0{target.username}\u00a0» réinitialisé avec succès.")
            return redirect("attendance:user_list")
    else:
        form = UserResetPasswordForm()

    return render(request, "attendance/user_reset_password.html", {
        "form": form,
        "target_user": target,
    })


# ─────────────────────────────────────────────────────────────────────────────
# JOURNAL D'ACTIVITE SYSTEME
# ─────────────────────────────────────────────────────────────────────────────

def system_log_view(request: HttpRequest) -> HttpResponse:
    """Journal d'activité système — filtres + pagination."""
    guard = _require_admin(request)
    if guard:
        return guard

    qs = SystemLog.objects.order_by("-created_at")

    # ── Filtres ───────────────────────────────────────────────────────────────
    action_filter = request.GET.get("action", "")
    obj_filter = request.GET.get("obj", "")
    success_filter = request.GET.get("success", "")
    user_filter = request.GET.get("user", "").strip()

    if action_filter:
        qs = qs.filter(action=action_filter)
    if obj_filter:
        qs = qs.filter(object_type=obj_filter)
    if success_filter == "1":
        qs = qs.filter(success=True)
    elif success_filter == "0":
        qs = qs.filter(success=False)
    if user_filter:
        qs = qs.filter(user__icontains=user_filter)

    total = qs.count()
    page_obj = Paginator(qs, 50).get_page(request.GET.get("page", 1))

    # ── KPIs 7 derniers jours ─────────────────────────────────────────────────
    from django.utils import timezone as tz_util
    depuis_7j = tz_util.now() - timedelta(days=7)
    kpi = {
        "logins": SystemLog.objects.filter(
            action=SystemLog.ACTION_LOGIN, created_at__gte=depuis_7j, success=True).count(),
        "creates": SystemLog.objects.filter(
            action=SystemLog.ACTION_CREATE, created_at__gte=depuis_7j).count(),
        "trains": SystemLog.objects.filter(
            action=SystemLog.ACTION_TRAIN, created_at__gte=depuis_7j).count(),
        "errors": SystemLog.objects.filter(
            success=False, created_at__gte=depuis_7j).count(),
    }

    return render(request, "attendance/system_log.html", {
        "logs": page_obj,
        "page_obj": page_obj,
        "total": total,
        "kpi": kpi,
        "action_choices": SystemLog.ACTION_CHOICES,
        "obj_choices": SystemLog.OBJ_CHOICES,
        "action_filter": action_filter,
        "obj_filter": obj_filter,
        "success_filter": success_filter,
        "user_filter": user_filter,
    })
